#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
主界面 - 通道查询、编辑、批量修改、Excel导入导出
v7.0 - 新增设备列表选择，登录后自动查询全部国标设备
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog
import hashlib
import requests
import threading
from datetime import datetime
import os
from concurrent.futures import ThreadPoolExecutor, as_completed

# 尝试导入 openpyxl（非必需，但导出/导入需要）
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from progress_dialog import ProgressDialog


class MainApplication:
    def __init__(self, root, token, user_info, host):
        self.root = root
        self.access_token = token
        self.login_user = user_info
        self.server_host = tk.StringVar(value=host)

        self.root.title("WVP通道管理工具 v7.0 by Xiaoabiao")
        self.root.geometry("1200x800")
        self.root.minsize(1000, 650)
        self.root.resizable(True, True)

        # 设备列表相关
        self.all_devices = []           # 全部设备列表
        self.selected_device_id = None  # 当前选中的设备ID
        self.selected_device_name = ""
        self.selected_device_status = ""

        self.status_text = tk.StringVar(value=f"已登录: {user_info.get('username', '')}")
        self.page_num = 1
        self.page_size = 50000
        self.total_channels = 0
        self.all_channels = []
        self.item_to_channel = {}

        self.edit_entry = None
        self.edit_item = None
        self.edit_column = None

        self.check_vars = {}
        self.select_all_var = tk.BooleanVar(value=False)

        # 设备列表复选框
        self.dev_check_vars = {}

        self.progress = None          # 进度弹窗实例
        self.max_workers = 8          # 并发线程数（可自定义）

        self.setup_styles()
        self.build_ui()
        # 登录后自动查询设备列表
        self.root.after(500, self.do_query_devices)
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def setup_styles(self):
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("Treeview", rowheight=26, font=("Microsoft YaHei", 9))
        style.configure("Treeview.Heading", font=("Microsoft YaHei", 10, "bold"))
        style.configure("TButton", font=("Microsoft YaHei", 9))
        style.configure("TLabel", font=("Microsoft YaHei", 9))
        style.configure("TEntry", font=("Microsoft YaHei", 9))

    def build_ui(self):
        # 顶栏
        top_bar = ttk.Frame(self.root)
        top_bar.pack(fill=tk.X, padx=10, pady=(5, 0))
        ttk.Label(top_bar, text=f"  {self.login_user.get('username', '')}",
                  font=("Microsoft YaHei", 10, "bold")).pack(side=tk.LEFT)
        self.thread_btn = ttk.Button(top_bar, text=f"并发: {self.max_workers}",
                                      command=self.set_max_workers, width=10)
        self.thread_btn.pack(side=tk.RIGHT, padx=5)
        ttk.Button(top_bar, text="退出登录", command=self.logout).pack(side=tk.RIGHT, padx=5)
        ttk.Separator(self.root, orient='horizontal').pack(fill=tk.X, padx=10, pady=(2, 0))

        # ========== 主内容区域：左侧设备列表 + 右侧通道列表 ==========
        main_pw = ttk.PanedWindow(self.root, orient=tk.HORIZONTAL)
        main_pw.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

        # ----- 左侧：设备列表 -----
        left_frame = ttk.LabelFrame(main_pw, text=" 国标设备列表 ", padding=5)
        main_pw.add(left_frame, weight=1)  # 较小的权重

        # 设备列表顶栏（按钮横排）
        dev_toolbar = ttk.Frame(left_frame)
        dev_toolbar.pack(fill=tk.X, pady=(0, 5))
        ttk.Button(dev_toolbar, text=" 刷新设备", command=self.do_query_devices, width=12).pack(side=tk.LEFT, padx=(0, 3))
        ttk.Button(dev_toolbar, text=" 导出勾选通道", command=self.export_selected_devices, width=14).pack(side=tk.LEFT, padx=3)
        ttk.Button(dev_toolbar, text=" 导入通道Excel", command=self.import_device_excel, width=14).pack(side=tk.LEFT, padx=3)
        ttk.Button(dev_toolbar, text=" 一键导入文件夹", command=self.batch_import_by_folder, width=16).pack(side=tk.LEFT, padx=3)
        ttk.Button(dev_toolbar, text=" 导入模板", command=self.download_import_template, width=10).pack(side=tk.LEFT, padx=3)
        self.dev_count_var = tk.StringVar(value="设备数: -")
        ttk.Label(dev_toolbar, textvariable=self.dev_count_var, foreground="gray").pack(side=tk.RIGHT, padx=5)

        # 设备列表（带复选框 + 状态/型号/厂家）
        dev_tree_frame = ttk.Frame(left_frame)
        dev_tree_frame.pack(fill=tk.BOTH, expand=True)
        dev_tree_frame.grid_rowconfigure(0, weight=1)
        dev_tree_frame.grid_columnconfigure(0, weight=1)

        dev_columns = ("☐", "设备ID", "名称", "在线", "型号", "厂家")
        self.dev_tree = ttk.Treeview(dev_tree_frame, columns=dev_columns, show="headings",
                                      selectmode="browse", height=15)
        dev_col_widths = {"☐": 30, "设备ID": 160, "名称": 120, "在线": 45, "型号": 80, "厂家": 80}
        for col in dev_columns:
            w = dev_col_widths.get(col, 80)
            self.dev_tree.heading(col, text=col)
            self.dev_tree.column(col, width=w, anchor=tk.CENTER, minwidth=w)
        # 点击 ☐ 表头 = 全选/取消全选
        self.dev_tree.heading("#1", command=self.toggle_dev_select_all)

        dev_vsb = ttk.Scrollbar(dev_tree_frame, orient=tk.VERTICAL, command=self.dev_tree.yview)
        self.dev_tree.configure(yscrollcommand=dev_vsb.set)
        self.dev_tree.grid(row=0, column=0, sticky="nsew")
        dev_vsb.grid(row=0, column=1, sticky="ns")

        self.dev_tree.bind("<Double-1>", self.on_device_double_click)
        self.dev_tree.bind("<Button-1>", self.on_dev_checkbox_click)

        # 底部状态栏
        self.statusbar_var = tk.StringVar(value="就绪")
        statusbar = ttk.Label(self.root, textvariable=self.statusbar_var, relief=tk.SUNKEN,
                              anchor=tk.W, font=("Microsoft YaHei", 8), padding=(5, 2))
        statusbar.pack(side=tk.BOTTOM, fill=tk.X)

        self.update_ui_state()

    # ---------- UI 状态 ----------
    def update_ui_state(self):
        # UI已精简，暂无动态控件需要更新
        pass

    def set_statusbar(self, msg):
        self.statusbar_var.set(f"{msg}  [{datetime.now().strftime('%H:%M:%S')}]")

    def _show_progress(self, title, total):
        """显示进度弹窗"""
        if self.progress:
            try:
                self.progress.close()
            except:
                pass
        self.progress = ProgressDialog(self.root, title, total)

    def _close_progress(self):
        if self.progress:
            try:
                self.progress.close()
            except:
                pass
            self.progress = None

    def set_max_workers(self):
        """弹窗修改并发线程数"""
        val = simpledialog.askinteger("并发设置",
            f"当前并发线程数: {self.max_workers}\n请输入新的数值（1-32）:",
            initialvalue=self.max_workers, minvalue=1, maxvalue=32, parent=self.root)
        if val:
            self.max_workers = val
            self.thread_btn.configure(text=f"并发: {self.max_workers}")
            self.set_statusbar(f"并发线程数已设为 {self.max_workers}")

    # ---------- 并发分页查询通道 ----------
    def _concurrent_channel_query(self, device_id, page_size=100, progress=None):
        """并发分页查询设备通道，返回 (channels, total)
           progress: 可选回调 (current, total, text)
        """
        host = self.server_host.get().strip().rstrip('/')
        token = self.access_token
        headers = {"Accept": "*/*", "access-token": token}
        url = f"{host}/api/device/query/devices/{device_id}/channels"
        base_params = {"query": "", "cameraQuery": "", "nvrQuery": ""}

        if progress:
            progress(0, 1, "正在获取总页数...")

        # 先查第一页获取总数
        try:
            r = requests.get(url, headers=headers, params={
                **base_params, "page": 1, "count": page_size}, timeout=30)
            if r.status_code != 200:
                return [], 0
            data = r.json()
            if data.get("code") != 0:
                return [], 0
            inner = data.get("data", {})
            first_batch = inner.get("list", [])
            total = inner.get("total", 0) or len(first_batch)
        except Exception:
            return [], 0

        if total <= page_size:
            if progress:
                progress(1, 1, f"查询完成，共 {total} 个通道")
            return first_batch, total

        total_pages = (total + page_size - 1) // page_size
        pages_list = {1: first_batch}
        page_range = range(2, total_pages + 1)

        if progress:
            progress(0, total_pages, f"共 {total_pages} 页，正在并发查询...")

        def fetch_page(pg):
            try:
                resp = requests.get(url, headers=headers, params={
                    **base_params, "page": pg, "count": page_size}, timeout=30)
                if resp.status_code == 200:
                    d = resp.json()
                    if d.get("code") == 0:
                        return pg, d.get("data", {}).get("list", [])
            except Exception:
                pass
            return pg, []

        fetched = [len(first_batch)]
        done = [1]
        with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
            fut_map = {ex.submit(fetch_page, p): p for p in page_range}
            for fut in as_completed(fut_map):
                pg, lst = fut.result()
                pages_list[pg] = lst
                fetched[0] += len(lst)
                done[0] += 1
                if progress:
                    progress(fetched[0], total,
                        f"正在查询 ({fetched[0]}/{total} 个通道)...")

        all_channels = []
        for pg in sorted(pages_list):
            all_channels.extend(pages_list[pg])

        if progress:
            progress(total, total, f"查询完成，共 {total} 个通道")
        return all_channels, total

    # ========== 设备列表相关 ==========

    def do_query_devices(self):
        """查询全部国标设备"""
        if not self.access_token:
            messagebox.showwarning("警告", "请先登录")
            return
        self.set_statusbar("正在查询全部国标设备...")
        self.dev_count_var.set("查询中...")
        host = self.server_host.get().strip().rstrip('/')
        token = self.access_token

        def task():
            nonlocal host, token
            try:
                url = f"{host}/api/device/query/devices"
                headers = {
                    "Accept": "*/*",
                    "access-token": token,
                    "Content-Type": "application/x-www-form-urlencoded"
                }
                all_devices = []
                page = 1
                count = 100
                total = 0

                while True:
                    params = {
                        "page": page,
                        "count": count,
                        "query": "",
                        "status": ""
                    }
                    resp = requests.get(url, headers=headers, params=params, timeout=30)

                    if resp.status_code == 200:
                        data = resp.json()
                        if data.get("code") != 0:
                            self.root.after(0, lambda m=data.get("msg", "未知错误"): self._devices_fail(m))
                            return

                        inner = data.get("data", {})
                        devices = inner.get("list", [])
                        total = inner.get("total", 0)
                        all_devices.extend(devices)

                        self.root.after(0, lambda p=page: self.set_statusbar(f"正在加载设备第 {p} 页..."))

                        if len(devices) < count or len(all_devices) >= total:
                            break
                        page += 1
                    else:
                        msg = resp.json().get("msg", f"HTTP {resp.status_code}")
                        self.root.after(0, lambda m=msg: self._devices_fail(m))
                        return

                self.all_devices = all_devices
                self.root.after(0, lambda: self._devices_success(all_devices, total))

            except Exception as e:
                self.root.after(0, lambda e=e: self._devices_fail(str(e)))

        threading.Thread(target=task, daemon=True).start()

    def _devices_success(self, devices, total):
        self.dev_tree.delete(*self.dev_tree.get_children())
        self.dev_check_vars.clear()
        self.dev_count_var.set(f"设备数: {total}")

        for dev in devices:
            device_id = dev.get("deviceId", "")
            name = dev.get("name", "")
            online = "ON" if dev.get("onLine") else "OFF"
            model = dev.get("model", dev.get("gbModel", ""))
            mfr = dev.get("manufacturer", dev.get("gbManufacturer", ""))
            var = tk.BooleanVar(value=False)
            values = ("☐", device_id, name, online, model, mfr)
            item = self.dev_tree.insert("", tk.END, values=values)
            self.dev_check_vars[item] = var

        self.set_statusbar(f"查询成功 - 共 {total} 个国标设备")
        self.update_ui_state()

    def _update_dev_header(self):
        """更新设备列表 ☐ 表头状态"""
        if not self.dev_check_vars:
            return
        all_checked = all(var.get() for var in self.dev_check_vars.values())
        any_checked = any(var.get() for var in self.dev_check_vars.values())
        if all_checked:
            self.dev_tree.heading("#1", text="☑")
        elif any_checked:
            self.dev_tree.heading("#1", text="☐")
        else:
            self.dev_tree.heading("#1", text="☐")

    def toggle_dev_select_all(self):
        """点击 ☐ 表头 = 全选/取消全选"""
        if not self.dev_check_vars:
            return
        # 如果全部已勾选则取消全选，否则全选
        all_checked = all(var.get() for var in self.dev_check_vars.values())
        new_state = not all_checked
        for item, var in self.dev_check_vars.items():
            var.set(new_state)
            self.dev_tree.set(item, "#1", "☑" if new_state else "☐")
        self._update_dev_header()

    def on_dev_checkbox_click(self, event):
        """设备列表复选框点击（仅处理数据行的点击，表头点击由 toggle_dev_select_all 处理）"""
        region = self.dev_tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        col = self.dev_tree.identify_column(event.x)
        if col != "#1":
            return
        item = self.dev_tree.identify_row(event.y)
        if not item:
            return
        var = self.dev_check_vars.get(item)
        if var:
            var.set(not var.get())
            self.dev_tree.set(item, "#1", "☑" if var.get() else "☐")
            self._update_dev_header()
        return "break"

    def _devices_fail(self, msg):
        self.dev_count_var.set("查询失败")
        self.set_statusbar(f"设备列表查询失败: {msg}")
        messagebox.showerror("查询失败", msg)

    def on_device_double_click(self, event):
        """设备列表双击 - 弹出新窗口显示通道"""
        item = self.dev_tree.identify_row(event.y)
        if not item:
            return
        values = self.dev_tree.item(item, "values")
        if not values:
            return
        device_id = values[1]
        device_name = values[2]
        self.set_statusbar(f"正在加载设备 {device_name} 的通道...")
        threading.Thread(target=self._open_device_window,
                         args=(device_id, device_name), daemon=True).start()

    def _open_device_window(self, device_id, device_name):
        """后台查询通道并在新窗口展示（并发分页+进度条）"""
        host = self.server_host.get().strip().rstrip('/')
        token = self.access_token
        try:
            headers = {"Accept": "*/*", "access-token": token}
            self.root.after(0, lambda: self._show_progress(
                f"正在查询 {device_name}", 1))

            def on_progress(c, t, text):
                self.root.after(0, lambda: self.progress.update(c, t, text))

            channels, total = self._concurrent_channel_query(
                device_id, progress=on_progress)
            # 进度条继续显示合并全局通道数据
            merge_done = [0]
            merge_total = len(channels)

            def merge_one(dc):
                dc_id = dc.get("id") or dc.get("gbId")
                if not dc_id:
                    return dc
                try:
                    one_resp = requests.get(f"{host}/api/common/channel/one",
                                            headers=headers,
                                            params={"id": dc_id}, timeout=15)
                    if one_resp.status_code == 200:
                        od = one_resp.json()
                        if od.get("code") == 0:
                            gc = od.get("data")
                            if gc and isinstance(gc, dict):
                                for fld in ("gbManufacturer", "gbLongitude", "gbLatitude",
                                            "gbName", "gbCivilCode",
                                            "gbModel", "gbOwner", "gbAddress", "gbPassword", "gbDeviceId"):
                                    if fld in gc and gc[fld] is not None:
                                        dc[fld] = gc[fld]
                except Exception:
                    pass
                return dc

            def merge_with_progress(dc):
                result = merge_one(dc)
                merge_done[0] += 1
                if merge_done[0] % 50 == 0:
                    on_progress(merge_done[0], merge_total,
                        f"合并通道数据 ({merge_done[0]}/{merge_total})...")
                return result

            with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
                list(ex.map(merge_with_progress, channels))
            self.root.after(0, lambda: self._close_progress())
        except Exception as e:
            self.root.after(0, lambda e=e:
                messagebox.showerror("查询失败", str(e)))
            return
        # 在新窗口中显示
        self.root.after(0, lambda: self._show_device_window(
            device_id, device_name, channels))

    def _show_device_window(self, device_id, device_name, channels):
        """创建新窗口展示设备通道列表（含编辑、导入导出）"""
        win = tk.Toplevel(self.root)
        win.title(f"通道列表 - {device_name} ({device_id})")
        win.geometry("1000x680")
        win.minsize(850, 500)

        # ---- 顶部工具栏 ----
        toolbar = ttk.Frame(win)
        toolbar.pack(fill=tk.X, padx=10, pady=(8, 0))
        ttk.Label(toolbar, text=f"设备: {device_name}    ID: {device_id}",
                  font=("Microsoft YaHei", 11, "bold")).pack(side=tk.LEFT)
        ttk.Label(toolbar, text=f"共 {len(channels)} 个通道",
                  foreground="gray").pack(side=tk.RIGHT)

        # 按钮行
        btn_row = ttk.Frame(win)
        btn_row.pack(fill=tk.X, padx=10, pady=(5, 0))
        ttk.Button(btn_row, text=" 刷新",
                   command=lambda: self._win_refresh(win, device_id, device_name,
                                                     tree, item_to_ch, check_vars, set_st)).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="导出Excel",
                   command=lambda: self._win_export(win, channels, device_name)).pack(side=tk.LEFT, padx=(0, 5))
        ttk.Button(btn_row, text="导入Excel",
                   command=lambda: self._win_import(win, channels, device_id, device_name)).pack(side=tk.LEFT, padx=5)
        ttk.Button(btn_row, text="批量修改区域编码",
                   command=lambda: self._win_batch_region(win, tree, check_vars, item_to_ch)).pack(side=tk.LEFT, padx=5)
        ttk.Label(btn_row, text="双击单元格修改", foreground="gray",
                  font=("Microsoft YaHei", 9)).pack(side=tk.RIGHT)

        # ---- 通道表格（带复选框）----
        frame = ttk.Frame(win)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)
        frame.grid_rowconfigure(0, weight=1)
        frame.grid_columnconfigure(0, weight=1)

        columns = ("☐", "序号", "通道名称", "通道类型", "区域编码", "经度", "纬度", "厂家", "数据库ID")
        tree = ttk.Treeview(frame, columns=columns, show="headings", height=20)
        col_widths = [30, 40, 200, 80, 100, 80, 80, 120, 80]
        for col, w in zip(columns, col_widths):
            tree.heading(col, text=col)
            tree.column(col, width=w, anchor=tk.CENTER, minwidth=w)

        # 表头 ☐ 点击 = 全选/取消全选
        def update_ch_header():
            if not check_vars:
                return
            all_checked = all(v.get() for v in check_vars.values())
            tree.heading("#1", text="☑" if all_checked else "☐")

        def toggle_all_ch():
            if not check_vars:
                return
            all_checked = all(v.get() for v in check_vars.values())
            new_state = not all_checked
            for item, v in check_vars.items():
                v.set(new_state)
                tree.set(item, "#1", "☑" if new_state else "☐")
            update_ch_header()
        tree.heading("#1", command=toggle_all_ch)

        vsb = ttk.Scrollbar(frame, orient=tk.VERTICAL, command=tree.yview)
        hsb = ttk.Scrollbar(frame, orient=tk.HORIZONTAL, command=tree.xview)
        tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # 存储数据
        item_to_ch = {}
        check_vars = {}
        for i, ch in enumerate(channels, 1):
            var = tk.BooleanVar(value=False)
            item = tree.insert("", tk.END, values=(
                "☐", i,
                ch.get("name", ""),
                "子目录" if ch.get("channelType") else "设备通道",
                ch.get("civilCode", ""),
                ch.get("gbLongitude", 0),
                ch.get("gbLatitude", 0),
                ch.get("gbManufacturer", ""),
                ch.get("id", ""),
            ))
            item_to_ch[item] = ch
            check_vars[item] = var

        # 复选框点击
        def on_chk_click(event):
            region = tree.identify_region(event.x, event.y)
            if region != "cell":
                return
            col = tree.identify_column(event.x)
            if col != "#1":
                return
            it = tree.identify_row(event.y)
            if it and it in check_vars:
                v = check_vars[it]
                v.set(not v.get())
                tree.set(it, "#1", "☑" if v.get() else "☐")
                update_ch_header()
            return "break"

        tree.bind("<Button-1>", on_chk_click)

        # 双击编辑
        edit_entry_ref = [None]
        edit_item_ref = [None]
        edit_col_ref = [None]

        def cancel_edit():
            if edit_entry_ref[0]:
                edit_entry_ref[0].destroy()
                edit_entry_ref[0] = None
                edit_item_ref[0] = None
                edit_col_ref[0] = None

        def save_edit():
            if not edit_entry_ref[0]:
                return
            new_val = edit_entry_ref[0].get().strip()
            it, col = edit_item_ref[0], edit_col_ref[0]
            cancel_edit()
            if not it or not new_val:
                return
            ch = item_to_ch.get(it)
            if not ch:
                return
            # 判断是哪一列
            if col == "#3":  # 通道名称
                if new_val == ch.get("name", ""):
                    return
                upd = {"gbName": new_val}
                fld = "名称"
            elif col == "#5":  # 区域编码
                if new_val == ch.get("civilCode", ""):
                    return
                upd = {"gbCivilCode": new_val}
                fld = "区域编码"
            elif col == "#6":  # 经度
                try:
                    nv = float(new_val)
                except ValueError:
                    messagebox.showerror("错误", "经度必须是数字", parent=win)
                    return
                old = float(ch.get("gbLongitude", 0) or 0)
                if abs(nv - old) < 0.000001:
                    return
                upd = {"gbLongitude": nv}
                fld = "经度"
            elif col == "#7":  # 纬度
                try:
                    nv = float(new_val)
                except ValueError:
                    messagebox.showerror("错误", "纬度必须是数字", parent=win)
                    return
                old = float(ch.get("gbLatitude", 0) or 0)
                if abs(nv - old) < 0.000001:
                    return
                upd = {"gbLatitude": nv}
                fld = "纬度"
            elif col == "#8":  # 厂家
                if new_val == ch.get("gbManufacturer", ""):
                    return
                upd = {"gbManufacturer": new_val}
                fld = "厂家"
            else:
                return
            if not messagebox.askyesno("确认",
                    f"将 {fld} 改为 '{new_val}'？", parent=win):
                return
            # 发请求
            host = self.server_host.get().strip().rstrip('/')
            headers = {"Accept": "*/*", "access-token": self.access_token,
                       "Content-Type": "application/json"}
            body = self.build_channel_body(ch, upd)
            try:
                resp = requests.post(f"{host}/api/common/channel/update",
                                     headers=headers, json=body, timeout=15)
                if resp.status_code == 200 and resp.json().get("code") == 0:
                    tree.set(it, col, new_val)
                    for k, v in upd.items():
                        ch[k] = v
                    # 同步短名字段，确保后续比较一致
                    if "gbName" in upd:
                        ch["name"] = upd["gbName"]
                    if "gbCivilCode" in upd:
                        ch["civilCode"] = upd["gbCivilCode"]
                    self.set_statusbar(f"修改成功: {fld}")
                else:
                    msg = resp.json().get("msg", "修改失败")
                    messagebox.showerror("失败", msg, parent=win)
            except Exception as e:
                messagebox.showerror("异常", str(e), parent=win)

        def on_dblclick(event):
            col = tree.identify_column(event.x)
            if col in ("#1", "#2", "#4", "#9"):
                return
            cancel_edit()
            it = tree.identify_row(event.y)
            if not it:
                return
            val = tree.set(it, col)
            bbox = tree.bbox(it, col)
            if not bbox:
                return
            x, y, w, h = bbox
            entry = tk.Entry(tree, font=("Microsoft YaHei", 9))
            entry.place(x=x, y=y, width=w, height=h)
            entry.insert(0, val)
            entry.select_range(0, tk.END)
            entry.focus_set()
            entry.bind("<Return>", lambda e: save_edit())
            entry.bind("<FocusOut>", lambda e: win.after(200, save_edit))
            edit_entry_ref[0] = entry
            edit_item_ref[0] = it
            edit_col_ref[0] = col

        tree.bind("<Double-1>", on_dblclick)

        # 底部状态
        status_var = tk.StringVar(value="就绪")
        status_bar = ttk.Label(win, textvariable=status_var, relief=tk.SUNKEN,
                               anchor=tk.W, font=("Microsoft YaHei", 8), padding=(5, 2))
        status_bar.pack(side=tk.BOTTOM, fill=tk.X)

        def set_st(msg):
            status_var.set(msg)

        # ---- 窗口内的导入导出 ----
        def do_win_export():
            if not channels:
                messagebox.showwarning("提示", "通道列表为空", parent=win)
                return
            f = filedialog.asksaveasfilename(defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")], title="导出通道",
                initialfile=f"{device_name}.xlsx")
            if not f:
                return
            wb = Workbook()
            ws = wb.active
            ws.title = "通道列表"
            hdrs = ["设备名称", "设备ID", "通道名称", "通道类型", "区域编码",
                    "经度", "纬度", "厂家", "数据库ID",
                    "设备型号", "设备归属", "地址", "密码", "国标编码"]
            for c, h in enumerate(hdrs, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for i, ch in enumerate(channels, 1):
                row = [device_name, device_id,
                       ch.get("name", ""),
                       "子目录" if ch.get("channelType") else "设备通道",
                       ch.get("civilCode", ""),
                       ch.get("gbLongitude", 0), ch.get("gbLatitude", 0),
                       ch.get("gbManufacturer", ""), ch.get("id", ""),
                       ch.get("gbModel", ""), ch.get("gbOwner", ""),
                       ch.get("gbAddress", ""), ch.get("gbPassword", ""),
                       ch.get("gbDeviceId", "")]
                for j, v in enumerate(row, 1):
                    ws.cell(row=i + 1, column=j, value=v)
            for i, w in enumerate([18, 24, 25, 10, 15, 10, 10, 15, 10, 10, 12, 18, 10, 10], 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = w
            wb.save(f)
            set_st(f"导出成功: {os.path.basename(f)}")

        def do_win_import():
            f = filedialog.askopenfilename(title="选择修改后的Excel",
                filetypes=[("Excel", "*.xlsx")])
            if not f:
                return
            try:
                wb = load_workbook(f)
                ws = wb.active
            except Exception as e:
                messagebox.showerror("读取失败", str(e), parent=win)
                return
            excel_data = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 9:
                    continue
                did = row[8]
                if not did:
                    continue
                try:
                    did = int(did)
                except:
                    continue
                excel_data[did] = {
                    "name": str(row[2] or "").strip(),
                    "civilCode": str(row[4] or "").strip(),
                    "gbLongitude": str(row[5] or "").strip(),
                    "gbLatitude": str(row[6] or "").strip(),
                    "gbManufacturer": str(row[7] or "").strip(),
                    "gbModel": str(row[9] if len(row) > 9 else "" or "").strip(),
                    "gbOwner": str(row[10] if len(row) > 10 else "" or "").strip(),
                    "gbAddress": str(row[11] if len(row) > 11 else "" or "").strip(),
                    "gbPassword": str(row[12] if len(row) > 12 else "" or "").strip(),
                }
            if not excel_data:
                messagebox.showwarning("无数据", "Excel无有效数据", parent=win)
                return
            tasks = []
            for ch in channels:
                cid = ch.get("id")
                if cid in excel_data:
                    d = excel_data[cid]
                    n = d["name"]
                    c = d["civilCode"]
                    lo = d["gbLongitude"]
                    la = d["gbLatitude"]
                    mf = d["gbManufacturer"]
                    md = d.get("gbModel", "")
                    ow = d.get("gbOwner", "")
                    ad = d.get("gbAddress", "")
                    pw = d.get("gbPassword", "")
                    if ((n and n != ch.get("name", "")) or (c and c != ch.get("civilCode", "")) or
                        (lo and str(lo) != str(ch.get("gbLongitude", 0))) or
                        (la and str(la) != str(ch.get("gbLatitude", 0))) or
                        mf != ch.get("gbManufacturer", "") or
                        (md and md != ch.get("gbModel", "")) or
                        (ow and ow != ch.get("gbOwner", "")) or
                        (ad and ad != ch.get("gbAddress", "")) or
                        (pw and pw != ch.get("gbPassword", ""))):
                        tasks.append((ch, n, c, lo, la, mf, md, ow, ad, pw))
            if not tasks:
                messagebox.showinfo("提示", "没有需要修改的数据", parent=win)
                return
            if not messagebox.askyesno("确认",
                    f"检测到 {len(tasks)} 条修改，是否提交？", parent=win):
                return
            set_st("正在导入...")
            threading.Thread(target=self._win_batch_update,
                args=(win, tasks, tree, item_to_ch, set_st), daemon=True).start()

        # 绑定按钮
        for child in btn_row.winfo_children():
            if isinstance(child, ttk.Button):
                txt = child.cget("text")
                if "导出" in txt:
                    child.configure(command=do_win_export)
                elif "导入" in txt:
                    child.configure(command=do_win_import)

        # 关闭按钮
        ttk.Button(win, text="关闭", command=win.destroy).pack(pady=(0, 8))

    def _win_batch_update(self, win, tasks, tree, item_to_ch, set_st):
        """窗口内批量提交导入的修改（并发8线程）"""
        host = self.server_host.get().strip().rstrip('/')
        headers = {"Accept": "*/*", "access-token": self.access_token,
                   "Content-Type": "application/json"}
        total = len(tasks)
        done_count = [0]
        success = [0]
        fail = [0]
        updated_channels = []

        def do_update(task):
            ch, nn, nc, lo, la, mf, md, ow, ad, pw = task
            upd = {}
            # 对比时优先用 gb 前缀字段（合并后的值）
            if nn and nn != ch.get("gbName", ch.get("name", "")):
                upd["gbName"] = nn
            if nc and nc != ch.get("gbCivilCode", ch.get("civilCode", "")):
                upd["gbCivilCode"] = nc
            if lo:
                try:
                    val = float(lo)
                    if abs(val - float(ch.get("gbLongitude", 0) or 0)) > 0.000001:
                        upd["gbLongitude"] = val
                except ValueError:
                    pass
            if la:
                try:
                    val = float(la)
                    if abs(val - float(ch.get("gbLatitude", 0) or 0)) > 0.000001:
                        upd["gbLatitude"] = val
                except ValueError:
                    pass
            if mf and mf != ch.get("gbManufacturer", ""):
                upd["gbManufacturer"] = mf
            if md and md != ch.get("gbModel", ""):
                upd["gbModel"] = md
            if ow and ow != ch.get("gbOwner", ""):
                upd["gbOwner"] = ow
            if ad and ad != ch.get("gbAddress", ""):
                upd["gbAddress"] = ad
            if pw and pw != ch.get("gbPassword", ""):
                upd["gbPassword"] = pw
            if not upd:
                return True, ch, None
            try:
                body = self.build_channel_body(ch, upd)
                resp = requests.post(f"{host}/api/common/channel/update",
                                     headers=headers, json=body, timeout=15)
                if resp.status_code == 200 and resp.json().get("code") == 0:
                    for k, v in upd.items():
                        ch[k] = v
                    # 同步更新短名字段，确保 tree 显示刷新
                    if "gbName" in upd:
                        ch["name"] = upd["gbName"]
                    if "gbCivilCode" in upd:
                        ch["civilCode"] = upd["gbCivilCode"]
                    return True, ch, None
            except Exception:
                pass
            return False, None, upd

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(do_update, task): task for task in tasks}
            for future in as_completed(futures):
                ok, ch, upd = future.result()
                if ok:
                    success[0] += 1
                    if ch:
                        updated_channels.append(ch)
                else:
                    fail[0] += 1
                done_count[0] += 1
                self.root.after(0, lambda c=done_count[0], t=total:
                    set_st(f"导入中 ({c}/{t})..."))
        self.root.after(0, lambda: set_st(f"导入完成: 成功 {success[0]}, 失败 {fail[0]}"))
        self.root.after(0, lambda: messagebox.showinfo("结果",
            f"成功: {success[0]} 条\n失败: {fail[0]} 条", parent=win))
        if success[0] > 0:
            self.root.after(0, lambda: self._refresh_win_tree(tree, [t[0] for t in tasks], item_to_ch))

    def _win_batch_region(self, win, tree, check_vars, item_to_ch):
        """弹窗内批量修改区域编码"""
        checked = [it for it, var in check_vars.items() if var.get()]
        if not checked:
            messagebox.showwarning("提示", "请先勾选要修改的通道", parent=win)
            return
        new_region = simpledialog.askstring("批量修改区域编码",
                                            f"已选中 {len(checked)} 个通道\n请输入新的区域编码:",
                                            parent=win)
        if not new_region or not new_region.strip():
            return
        new_region = new_region.strip()
        if not messagebox.askyesno("确认",
                f"将为选中的 {len(checked)} 个通道设置区域编码为:\n'{new_region}'?\n\n确认修改?",
                parent=win):
            return
        threading.Thread(target=self._win_do_batch_region,
                         args=(win, checked, tree, new_region, item_to_ch), daemon=True).start()

    def _win_do_batch_region(self, win, checked_items, tree, new_region, item_to_ch):
        """实际执行批量修改（并发8线程）"""
        host = self.server_host.get().strip().rstrip('/')
        headers = {"Accept": "*/*", "access-token": self.access_token,
                   "Content-Type": "application/json"}
        total = len(checked_items)
        done_count = [0]
        success = [0]
        fail = [0]

        def do_update(it):
            ch = item_to_ch.get(it)
            if not ch:
                return False, None, None
            updates = {"gbCivilCode": new_region, "civilCode": new_region}
            try:
                body = self.build_channel_body(ch, updates)
                resp = requests.post(f"{host}/api/common/channel/update",
                                     headers=headers, json=body, timeout=15)
                if resp.status_code == 200 and resp.json().get("code") == 0:
                    return True, ch, it
            except Exception:
                pass
            return False, None, None

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(do_update, it): it for it in checked_items}
            for future in as_completed(futures):
                ok, ch, it = future.result()
                if ok:
                    success[0] += 1
                    if ch and it:
                        ch["civilCode"] = new_region
                        self.root.after(0, lambda i=it: tree.set(i, "#5", new_region))
                else:
                    fail[0] += 1
                done_count[0] += 1
                self.root.after(0, lambda c=done_count[0], t=total:
                    self.set_statusbar(f"批量修改中 ({c}/{t})..."))
        self.root.after(0, lambda: self.set_statusbar(
            f"批量修改完成: 成功 {success[0]}, 失败 {fail[0]}"))
        self.root.after(0, lambda: messagebox.showinfo("结果",
            f"成功: {success[0]} 条\n失败: {fail[0]} 条", parent=win))

    def _win_refresh(self, win, device_id, device_name,
                     tree, item_to_ch, check_vars, set_st):
        """重新查询设备通道并刷新窗口表格"""
        set_st("刷新中...")
        threading.Thread(target=self._do_win_refresh,
            args=(win, device_id, device_name,
                  tree, item_to_ch, check_vars, set_st), daemon=True).start()

    def _do_win_refresh(self, win, device_id, device_name,
                        tree, item_to_ch, check_vars, set_st):
        """后台重新查询并刷新表格（并发分页）"""
        host = self.server_host.get().strip().rstrip('/')
        token = self.access_token
        try:
            headers = {"Accept": "*/*", "access-token": token}
            new_channels, total = self._concurrent_channel_query(device_id)
            # 并发合并全局通道数据
            merge_done = [0]
            merge_total = len(new_channels)

            def merge_one(dc):
                dc_id = dc.get("id") or dc.get("gbId")
                if not dc_id:
                    return dc
                try:
                    one_resp = requests.get(f"{host}/api/common/channel/one",
                                            headers=headers,
                                            params={"id": dc_id}, timeout=15)
                    if one_resp.status_code == 200:
                        od = one_resp.json()
                        if od.get("code") == 0:
                            gc = od.get("data")
                            if gc and isinstance(gc, dict):
                                for fld in ("gbManufacturer", "gbLongitude", "gbLatitude",
                                            "gbName", "gbCivilCode",
                                            "gbModel", "gbOwner", "gbAddress", "gbPassword", "gbDeviceId"):
                                    if fld in gc and gc[fld] is not None:
                                        dc[fld] = gc[fld]
                except Exception:
                    pass
                return dc

            def merge_with_progress(dc):
                result = merge_one(dc)
                merge_done[0] += 1
                if merge_done[0] % 100 == 0:
                    self.root.after(0, lambda c=merge_done[0], t=merge_total:
                        set_st(f"合并通道数据 ({c}/{t})..."))
                return result

            with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
                list(ex.map(merge_with_progress, new_channels))
        except Exception as e:
            self.root.after(0, lambda e=e: set_st(f"刷新失败: {e}"))
            return
        # 在主线程更新表格
        self.root.after(0, lambda: self._rebuild_win_tree(
            win, tree, new_channels, item_to_ch, check_vars, set_st))

    def _rebuild_win_tree(self, win, tree, new_channels, item_to_ch, check_vars, set_st):
        """重建窗口表格数据"""
        # 清除旧数据
        for item in tree.get_children():
            tree.delete(item)
        item_to_ch.clear()
        check_vars.clear()
        # 填入新数据
        for i, ch in enumerate(new_channels, 1):
            var = tk.BooleanVar(value=False)
            item = tree.insert("", tk.END, values=(
                "☐", i,
                ch.get("name", ""),
                "子目录" if ch.get("channelType") else "设备通道",
                ch.get("civilCode", ""),
                ch.get("gbLongitude", 0),
                ch.get("gbLatitude", 0),
                ch.get("gbManufacturer", ""),
                ch.get("id", ""),
            ))
            item_to_ch[item] = ch
            check_vars[item] = var
        # 更新标题栏的通道数
        for child in win.winfo_children():
            if isinstance(child, ttk.Frame):
                for sub in child.winfo_children():
                    if isinstance(sub, ttk.Label) and "共" in sub.cget("text"):
                        sub.configure(text=f"共 {len(new_channels)} 个通道")
                        break
        set_st(f"已刷新，共 {len(new_channels)} 个通道")

    @staticmethod
    def _refresh_win_tree(tree, channels, item_to_ch):
        """刷新窗口内表格显示"""
        for item in tree.get_children():
            ch = item_to_ch.get(item)
            if ch:
                tree.set(item, "#3", ch.get("gbName", ch.get("name", "")))
                tree.set(item, "#5", ch.get("gbCivilCode", ch.get("civilCode", "")))
                tree.set(item, "#6", ch.get("gbLongitude", 0))
                tree.set(item, "#7", ch.get("gbLatitude", 0))
                tree.set(item, "#8", ch.get("gbManufacturer", ""))

    # ========== 通道查询 ==========

    def do_query_channels(self):
        if not self.access_token:
            messagebox.showwarning("警告", "请先登录")
            return
        if not self.selected_device_id:
            messagebox.showwarning("提示", "请先从左侧设备列表中选择一个设备")
            return
        device_id = self.selected_device_id
        self.set_statusbar(f"正在查询设备 {device_id} 的全部通道...")
        self.query_btn.configure(state=tk.DISABLED, text="查询中...")
        self.cancel_edit()

        def task():
            try:
                host = self.server_host.get().strip().rstrip('/')
                headers = {
                    "Accept": "*/*",
                    "access-token": self.access_token,
                    "Content-Type": "application/x-www-form-urlencoded"
                }
                all_channels, total = self._concurrent_channel_query(device_id)

                if not all_channels and total == 0:
                    self.root.after(0, lambda: self._query_fail("未查询到通道数据"))
                    return

                # 并发查询全局通道，合并厂家/经纬度等字段
                try:
                    def merge_one(dc):
                        dc_id = dc.get("id") or dc.get("gbId")
                        if not dc_id:
                            return dc
                        try:
                            one_resp = requests.get(f"{host}/api/common/channel/one",
                                                    headers=headers,
                                                    params={"id": dc_id}, timeout=15)
                            if one_resp.status_code == 200:
                                od = one_resp.json()
                                if od.get("code") == 0:
                                    gc = od.get("data")
                                    if gc and isinstance(gc, dict):
                                        for fld in ("gbManufacturer", "gbLongitude", "gbLatitude",
                                                    "gbName", "gbCivilCode",
                                                    "gbModel", "gbOwner", "gbAddress", "gbPassword", "gbDeviceId"):
                                            if fld in gc and gc[fld] is not None:
                                                dc[fld] = gc[fld]
                        except Exception:
                            pass
                        return dc

                    with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
                        list(ex.map(merge_one, all_channels))
                except Exception as e:
                    import traceback
                    print(f"[DEBUG] 查询全局通道异常: {e}")
                    traceback.print_exc()

                self.all_channels = all_channels
                self.total_channels = total
                self.root.after(0, lambda: self._query_success(all_channels, total, device_id))

            except Exception as e:
                self.root.after(0, lambda e=e: self._query_fail(str(e)))

        threading.Thread(target=task, daemon=True).start()

    def _query_success(self, channels, total, device_id):
        self.query_btn.configure(state=tk.NORMAL, text=" 查询通道")
        self.tree.delete(*self.tree.get_children())
        self.item_to_channel.clear()
        self.check_vars.clear()
        self.select_all_var.set(False)

        for i, ch in enumerate(channels, 1):
            db_id = ch.get("id")
            ch_type = ch.get("channelType", 0)
            type_text = "子目录" if ch_type else "设备通道"
            var = tk.BooleanVar(value=False)
            values = ("☐", i, ch.get("deviceId", ""), ch.get("name", ""), type_text, ch.get("civilCode", ""), ch.get("gbLongitude", 0), ch.get("gbLatitude", 0), ch.get("gbManufacturer", ""), ch.get("id", ""))
            item = self.tree.insert("", tk.END, values=values)
            self.item_to_channel[item] = ch
            self.check_vars[item] = var
            var.trace_add("write", lambda *args, it=item: self.update_check_display(it))

        self.page_info_var.set(f"共 {total} 条")
        self.set_statusbar(f"查询成功 - 设备 {device_id} 下共 {total} 个通道")
        self.update_ui_state()

    def _query_fail(self, msg):
        self.query_btn.configure(state=tk.NORMAL, text=" 查询通道")
        self.set_statusbar(f"查询失败: {msg}")
        messagebox.showerror("查询失败", msg)

    # ---------- 复选框 ----------
    def update_check_display(self, item):
        var = self.check_vars.get(item)
        if var:
            self.tree.set(item, "#1", "☑" if var.get() else "☐")

    def on_checkbox_click(self, event):
        if self.tree.identify_column(event.x) != "#1":
            return
        item = self.tree.identify_row(event.y)
        if not item:
            return
        var = self.check_vars.get(item)
        if var:
            var.set(not var.get())
            self.update_select_all_state()

    def toggle_select_all(self):
        state = self.select_all_var.get()
        for var in self.check_vars.values():
            var.set(state)

    def update_select_all_state(self):
        if not self.check_vars:
            self.select_all_var.set(False)
            return
        all_checked = all(v.get() for v in self.check_vars.values())
        self.select_all_var.set(all_checked)

    def get_selected_channels(self):
        return [self.item_to_channel[it] for it, var in self.check_vars.items()
                if var.get() and it in self.item_to_channel]

    # ---------- 批量修改区域编码 ----------
    def batch_modify_region(self):
        selected = self.get_selected_channels()
        if not selected:
            messagebox.showwarning("提示", "请至少勾选一个通道")
            return
        new_region = simpledialog.askstring("批量修改区域编码",
                                            f"已选中 {len(selected)} 个通道\n请输入新的区域编码:",
                                            parent=self.root)
        if not new_region or not new_region.strip():
            return
        new_region = new_region.strip()
        if not messagebox.askyesno("确认", f"将为选中的 {len(selected)} 个通道设置区域编码为:\n'{new_region}'?\n\n确认修改?"):
            return
        self.batch_region_btn.configure(state=tk.DISABLED)
        self.set_statusbar("正在批量修改区域编码...")
        threading.Thread(target=self._batch_update_region, args=(selected, new_region), daemon=True).start()

    def _batch_update_region(self, channels, new_region):
        host = self.server_host.get().strip().rstrip('/')
        headers = {"Accept": "*/*", "access-token": self.access_token, "Content-Type": "application/json"}
        total = len(channels)
        done_count = [0]

        def do_update(ch):
            updates = {"gbCivilCode": new_region, "civilCode": new_region}
            try:
                body = self.build_channel_body(ch, updates)
                resp = requests.post(f"{host}/api/common/channel/update", headers=headers, json=body, timeout=15)
                if resp.status_code == 200 and resp.json().get("code") == 0:
                    return True
            except:
                pass
            return False

        success = 0
        fail = 0
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(do_update, ch): ch for ch in channels}
            for future in as_completed(futures):
                if future.result():
                    success += 1
                else:
                    fail += 1
                done_count[0] += 1
                self.root.after(0, lambda c=done_count[0], t=total:
                    self.set_statusbar(f"批量修改中: {c}/{t}"))
        self.root.after(0, lambda: self._batch_finished(success, fail))

    @staticmethod
    def _val(v, default=""):
        """取非None的值，None转默认值"""
        return v if v is not None else default

    @staticmethod
    def _num(v, default=0):
        """取非None的数值"""
        if v is None:
            return default
        try:
            return float(v)
        except (ValueError, TypeError):
            return default

    def build_channel_body(self, channel, updates):
        """构建 body：gbId + 有值的已有字段 + 要更新的字段
        注意：设备通道查询返回短名称(id, deviceId, name, civilCode)，
        但全局通道 API 需要 gb 前缀(gbId, gbDeviceId, gbName, gbCivilCode)，
        这里做字段映射。
        """
        # gbId：兼容设备通道(id)和全局通道(gbId)两种字段名
        gb_id_val = channel.get("id") or channel.get("gbId") or 0
        print(f"[DEBUG] build_channel_body: gbId={gb_id_val!r} (type={type(gb_id_val).__name__})")
        body = {"gbId": gb_id_val}
        # 字段映射：(body里的key, channel里的key, 默认值)
        field_map = [
            ("gbDeviceId", "deviceId", ""),
            ("gatewayDeviceId", "gatewayDeviceId", ""),
            ("gbName", "name", ""),
            ("gbManufacturer", "gbManufacturer", ""),
            ("gbModel", "gbModel", ""),
            ("gbOwner", "gbOwner", ""),
            ("gbCivilCode", "civilCode", ""),
            ("gbBlock", "gbBlock", ""),
            ("gbAddress", "gbAddress", ""),
            ("gbParental", "gbParental", 0),
            ("gbParentId", "gbParentId", ""),
            ("gbSafetyWay", "gbSafetyWay", 0),
            ("gbRegisterWay", "gbRegisterWay", 0),
            ("gbCertNum", "gbCertNum", ""),
            ("gbCertifiable", "gbCertifiable", 0),
            ("gbErrCode", "gbErrCode", 0),
            ("gbEndTime", "gbEndTime", ""),
            ("gbSecrecy", "gbSecrecy", 0),
            ("gbIpAddress", "gbIpAddress", ""),
            ("gbPort", "gbPort", 0),
            ("gbPassword", "gbPassword", ""),
            ("gbStatus", "gbStatus", ""),
            ("gbLongitude", "gbLongitude", 0),
            ("gbLatitude", "gbLatitude", 0),
            ("gpsAltitude", "gpsAltitude", 0),
            ("gpsSpeed", "gpsSpeed", 0),
            ("gpsDirection", "gpsDirection", 0),
            ("gpsTime", "gpsTime", ""),
            ("gbBusinessGroupId", "gbBusinessGroupId", ""),
            ("gbPtzType", "gbPtzType", 0),
            ("gbPositionType", "gbPositionType", 0),
            ("gbRoomType", "gbRoomType", 0),
            ("gbUseType", "gbUseType", 0),
            ("gbSupplyLightType", "gbSupplyLightType", 0),
            ("gbDirectionType", "gbDirectionType", 0),
            ("gbResolution", "gbResolution", ""),
            ("gbDownloadSpeed", "gbDownloadSpeed", ""),
            ("gbSvcSpaceSupportMod", "gbSvcSpaceSupportMod", 0),
            ("gbSvcTimeSupportMode", "gbSvcTimeSupportMode", 0),
            ("recordPLan", "recordPLan", 0),
            ("dataType", "dataType", 0),
            ("dataDeviceId", "dataDeviceId", 0),
            ("createTime", "createTime", ""),
            ("updateTime", "updateTime", ""),
        ]
        for body_key, ch_key, dflt in field_map:
            # 兼容短名称(id/name)和gb前缀(gbId/gbName)
            val = channel.get(ch_key) or channel.get(body_key) or dflt
            if val not in (None, "", 0, 0.0):
                body[body_key] = val
        body.update(updates)
        print(f"[DEBUG] build_channel_body 完成({len(body)}个字段): body={body}")
        return body

    def do_update(self, channel, updates, new_value, item, column):
        try:
            host = self.server_host.get().strip().rstrip('/')
            headers = {"Accept": "*/*", "access-token": self.access_token, "Content-Type": "application/json"}
            body = self.build_channel_body(channel, updates)
            url = f"{host}/api/common/channel/update"
            print(f"[DEBUG] ===== 开始修改通道 =====")
            print(f"[DEBUG] 请求URL: {url}")
            print(f"[DEBUG] 更新的字段: {updates}")
            print(f"[DEBUG] 完整body: {body}")
            resp = requests.post(url, headers=headers, json=body, timeout=15)
            print(f"[DEBUG] 响应状态码: {resp.status_code}")
            print(f"[DEBUG] 响应内容: {resp.text}")
            
            if resp.status_code == 200:
                resp_data = resp.json()
                if resp_data.get("code") == 0:
                    # 更新树显示
                    self.root.after(0, lambda: self.tree.set(item, column, new_value))
                    # 更新本地 channel dict
                    dc = self.item_to_channel.get(item)
                    if dc:
                        for k, v in updates.items():
                            dc[k] = v
                        # 同步短名字段，确保后续导入/导出比较一致
                        if "gbName" in updates:
                            dc["name"] = updates["gbName"]
                        if "gbCivilCode" in updates:
                            dc["civilCode"] = updates["gbCivilCode"]
                    self.root.after(0, lambda: self.set_statusbar("修改成功"))
                else:
                    msg = resp_data.get("msg", f"code={resp_data.get('code')}")
                    self.root.after(0, lambda m=msg: messagebox.showerror("修改失败", m))
            else:
                msg = f"HTTP {resp.status_code}"
                try:
                    msg = resp.json().get("msg", msg)
                except:
                    pass
                self.root.after(0, lambda m=msg: messagebox.showerror("修改失败", m))
        except Exception as e:
            import traceback
            print(f"[DEBUG] 修改异常: {e}")
            traceback.print_exc()
            self.root.after(0, lambda e=e: messagebox.showerror("修改异常", str(e)))

    def _batch_finished(self, success, fail):
        self.batch_region_btn.configure(state=tk.NORMAL)
        self.set_statusbar(f"批量修改完成: 成功 {success}, 失败 {fail}")
        messagebox.showinfo("结果", f"成功: {success}\n失败: {fail}")
        if success > 0:
            self.do_query_channels()

    # ---------- 双击编辑 ----------
    def on_double_click(self, event):
        col = self.tree.identify_column(event.x)
        if col == "#1":
            return
        if self.edit_entry:
            self.save_edit()
        item = self.tree.identify_row(event.y)
        if not item or col not in ("#4", "#6", "#7", "#8", "#9"):
            return
        self.edit_item = item
        self.edit_column = col
        value = self.tree.set(item, col)
        bbox = self.tree.bbox(item, col)
        if not bbox:
            return
        x, y, w, h = bbox
        self.edit_entry = tk.Entry(self.tree, font=("Microsoft YaHei", 9))
        self.edit_entry.place(x=x, y=y, width=w, height=h)
        self.edit_entry.insert(0, value)
        self.edit_entry.select_range(0, tk.END)
        self.edit_entry.focus_set()
        self.edit_entry.bind("<Return>", lambda e: self.save_edit())
        self.edit_entry.bind("<FocusOut>", self.on_focus_out)

    def on_focus_out(self, event):
        if self.edit_entry:
            self.root.after(100, self.save_edit)

    def save_edit(self):
        if not self.edit_entry:
            return
        new_value = self.edit_entry.get().strip()
        item, col = self.edit_item, self.edit_column
        self.cancel_edit()
        if not item or not new_value:
            return
        channel = self.item_to_channel.get(item)
        if not channel:
            return
        if col == "#4":
            old = channel.get("name", "")
            if new_value == old:
                return
            updates = {"gbName": new_value}
            field = "名称"
        elif col == "#6":
            old = channel.get("civilCode", "")
            if new_value == old:
                return
            updates = {"gbCivilCode": new_value}
            field = "区域编码"
        elif col == "#7":
            try:
                new_val = float(new_value)
            except ValueError:
                self.root.after(0, lambda: messagebox.showerror("错误", "经度必须是数字"))
                return
            old = channel.get("gbLongitude", 0)
            if old == "" or old is None:
                old = 0
            try:
                old = float(old)
            except (ValueError, TypeError):
                old = 0.0
            if abs(new_val - old) < 0.000001:
                return
            updates = {"gbLongitude": new_val}
            field = "经度"
        elif col == "#8":
            try:
                new_val = float(new_value)
            except ValueError:
                self.root.after(0, lambda: messagebox.showerror("错误", "纬度必须是数字"))
                return
            old = channel.get("gbLatitude", 0)
            if old == "" or old is None:
                old = 0
            try:
                old = float(old)
            except (ValueError, TypeError):
                old = 0.0
            if abs(new_val - old) < 0.000001:
                return
            updates = {"gbLatitude": new_val}
            field = "纬度"
        else:
            old = channel.get("gbManufacturer", "")
            if old is None:
                old = ""
            if new_value == old:
                return
            updates = {"gbManufacturer": new_value}
            field = "厂家"
        if not messagebox.askyesno("确认", f"将 {field} 从 '{old}' 改为 '{new_value}'?"):
            return
        threading.Thread(target=self.do_update, args=(channel, updates, new_value, item, col), daemon=True).start()

    def cancel_edit(self):
        if self.edit_entry:
            self.edit_entry.destroy()
            self.edit_entry = None
            self.edit_item = None
            self.edit_column = None



    # ---------- 导出选中设备的所有通道 ----------
    def export_selected_devices(self):
        """导出左侧勾选的设备下的全部通道"""
        if not self.access_token:
            messagebox.showwarning("警告", "请先登录")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("缺少库", "请安装 openpyxl: pip install openpyxl")
            return
        # 获取勾选的设备
        checked_items = [it for it, var in self.dev_check_vars.items() if var.get()]
        if not checked_items:
            messagebox.showwarning("提示", "请先在左侧设备列表中勾选要导出的设备（点击 ☐ 框）")
            return
        selected_devices = []
        for item in checked_items:
            values = self.dev_tree.item(item, "values")
            if values:
                dev_id = values[1]  # 设备ID在第2列（索引1）
                for dev in self.all_devices:
                    if dev.get("deviceId", "") == dev_id:
                        selected_devices.append(dev)
                        break
        if not selected_devices:
            messagebox.showwarning("提示", "未找到选中的设备数据")
            return
        if not messagebox.askyesno("确认", f"将导出 {len(selected_devices)} 个选中设备的全部通道，\n是否继续?"):
            return
        # 默认文件名 = 设备名称
        name_list = [d.get("name", f"设备{i+1}") for i, d in enumerate(selected_devices)]
        default_name = "、".join(name_list[:3])
        if len(name_list) > 3:
            default_name += f"等{len(name_list)}个设备"
        default_name += ".xlsx"
        file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx")],
                                             title="导出选中通道",
                                             initialfile=default_name)
        if not file:
            return
        self.set_statusbar("正在导出选中设备通道...")
        threading.Thread(target=self._do_export_selected,
                         args=(file, selected_devices), daemon=True).start()

    def _do_export_selected(self, file, devices):
        try:
            host = self.server_host.get().strip().rstrip('/')
            token = self.access_token
            headers = {"Accept": "*/*", "access-token": token}
            all_rows = []
            total = len(devices)
            for idx, dev in enumerate(devices, 1):
                device_id = dev.get("deviceId", "")
                device_name = dev.get("name", "")
                self.root.after(0, lambda c=idx, d=device_name:
                    self.set_statusbar(f"导出中 ({c}/{total}): {d}"))
                try:
                    channels, _ = self._concurrent_channel_query(device_id)
                    # 并发合并全局通道数据
                    def merge_one(dc):
                        dc_id = dc.get("id") or dc.get("gbId")
                        if not dc_id:
                            return dc
                        try:
                            one = requests.get(f"{host}/api/common/channel/one",
                                               headers=headers,
                                               params={"id": dc_id}, timeout=15)
                            if one.status_code == 200:
                                od = one.json()
                                if od.get("code") == 0:
                                    gc = od.get("data")
                                    if gc and isinstance(gc, dict):
                                        for f in ("gbManufacturer", "gbLongitude", "gbLatitude",
                                                  "gbName", "gbCivilCode",
                                                  "gbModel", "gbOwner", "gbAddress", "gbPassword", "gbDeviceId"):
                                            if f in gc and gc[f] is not None:
                                                dc[f] = gc[f]
                        except Exception:
                            pass
                        return dc
                    with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
                        list(ex.map(merge_one, channels))
                    for ch in channels:
                        all_rows.append((
                            device_name,
                            device_id,
                            ch.get("name", ""),
                            "子目录" if ch.get("channelType") else "设备通道",
                            ch.get("civilCode", ""),
                            ch.get("gbLongitude", 0),
                            ch.get("gbLatitude", 0),
                            ch.get("gbManufacturer", ""),
                            ch.get("id", ""),
                            ch.get("gbModel", ""),
                            ch.get("gbOwner", ""),
                            ch.get("gbAddress", ""),
                            ch.get("gbPassword", ""),
                            ch.get("gbDeviceId", ""),
                        ))
                except Exception:
                    pass
            # 写入 Excel
            wb = Workbook()
            ws = wb.active
            ws.title = "通道列表"
            headers_row = ["设备名称", "设备ID", "通道名称", "通道类型", "区域编码",
                           "经度", "纬度", "厂家", "数据库ID",
                           "设备型号", "设备归属", "地址", "密码", "国标编码"]
            for c, h in enumerate(headers_row, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for i, row_data in enumerate(all_rows, 1):
                for j, v in enumerate(row_data, 1):
                    ws.cell(row=i + 1, column=j, value=v)
            col_widths = [18, 24, 25, 10, 15, 10, 10, 15, 10, 10, 12, 18, 10, 10]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = w
            wb.save(file)
            self.root.after(0, lambda: self.set_statusbar(
                f"导出完成: {len(all_rows)} 条通道 → {os.path.basename(file)}"))
            self.root.after(0, lambda: messagebox.showinfo(
                "成功", f"共导出 {len(all_rows)} 条通道\n保存到: {file}"))
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.root.after(0, lambda e=e: messagebox.showerror("导出失败", str(e)))

    # ---------- 导入（从"导出勾选通道"生成的Excel）----------
    def import_device_excel(self):
        """导入修改后的通道Excel（支持多设备）"""
        if not self.access_token:
            messagebox.showwarning("警告", "请先登录")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("缺少库", "请安装 openpyxl: pip install openpyxl")
            return
        file = filedialog.askopenfilename(title="选择修改后的Excel",
                                           filetypes=[("Excel", "*.xlsx")])
        if not file:
            return
        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("读取失败", str(e))
            return
        # 读取每一行
        updates_by_id = {}  # db_id → {field: new_value}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or len(row) < 9:
                continue
            db_id = row[8]  # 数据库ID
            if not db_id:
                continue
            try:
                db_id = int(db_id)
            except (ValueError, TypeError):
                continue
            new_name = str(row[2] or "").strip()
            new_civil = str(row[4] or "").strip()
            new_lon = str(row[5] or "").strip()
            new_lat = str(row[6] or "").strip()
            new_mfr = str(row[7] or "").strip()
            new_model = str(row[9] if len(row) > 9 else "" or "").strip()
            new_owner = str(row[10] if len(row) > 10 else "" or "").strip()
            new_addr = str(row[11] if len(row) > 11 else "" or "").strip()
            new_pwd = str(row[12] if len(row) > 12 else "" or "").strip()
            updates_by_id[db_id] = (new_name, new_civil, new_lon, new_lat, new_mfr,
                                     new_model, new_owner, new_addr, new_pwd)
        if not updates_by_id:
            messagebox.showwarning("无数据", "Excel中未找到有效数据")
            return
        # 确认
        if not messagebox.askyesno("确认", f"检测到 {len(updates_by_id)} 条修改，是否提交?"):
            return
        self.set_statusbar("正在导入修改...")
        threading.Thread(target=self._do_import_device,
                         args=(file, updates_by_id), daemon=True).start()

    def _do_import_device(self, file, updates_by_id):
        """批量提交导入的修改（并发8线程）"""
        host = self.server_host.get().strip().rstrip('/')
        token = self.access_token
        headers = {"Accept": "*/*", "access-token": token,
                   "Content-Type": "application/json"}
        total = len(updates_by_id)
        done_count = [0]
        success = [0]
        fail = [0]

        def process_one(args):
            db_id, (new_name, new_civil, new_lon, new_lat, new_mfr,
                    new_model, new_owner, new_addr, new_pwd) = args
            try:
                one = requests.get(f"{host}/api/common/channel/one",
                                   headers=headers,
                                   params={"id": db_id}, timeout=15)
                if one.status_code != 200:
                    return False
                od = one.json()
                if od.get("code") != 0:
                    return False
                ch = od.get("data")
                if not ch or not isinstance(ch, dict):
                    return False
            except Exception:
                return False
            updates = {}
            if new_name and new_name != ch.get("gbName", ch.get("name", "")):
                updates["gbName"] = new_name
            if new_civil and new_civil != ch.get("gbCivilCode", ch.get("civilCode", "")):
                updates["gbCivilCode"] = new_civil
            if new_lon:
                try:
                    val = float(new_lon)
                    old = float(ch.get("gbLongitude", 0) or 0)
                    if abs(val - old) > 0.000001:
                        updates["gbLongitude"] = val
                except ValueError:
                    pass
            if new_lat:
                try:
                    val = float(new_lat)
                    old = float(ch.get("gbLatitude", 0) or 0)
                    if abs(val - old) > 0.000001:
                        updates["gbLatitude"] = val
                except ValueError:
                    pass
            if new_mfr and new_mfr != ch.get("gbManufacturer", ""):
                updates["gbManufacturer"] = new_mfr
            if new_model and new_model != ch.get("gbModel", ""):
                updates["gbModel"] = new_model
            if new_owner and new_owner != ch.get("gbOwner", ""):
                updates["gbOwner"] = new_owner
            if new_addr and new_addr != ch.get("gbAddress", ""):
                updates["gbAddress"] = new_addr
            if new_pwd and new_pwd != ch.get("gbPassword", ""):
                updates["gbPassword"] = new_pwd
            if not updates:
                return True
            try:
                body = self.build_channel_body(ch, updates)
                resp = requests.post(f"{host}/api/common/channel/update",
                                     headers=headers, json=body, timeout=15)
                if resp.status_code == 200 and resp.json().get("code") == 0:
                    return True
            except Exception:
                pass
            return False

        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            futures = {executor.submit(process_one, item): item
                       for item in updates_by_id.items()}
            for future in as_completed(futures):
                if future.result():
                    success[0] += 1
                else:
                    fail[0] += 1
                done_count[0] += 1
                self.root.after(0, lambda c=done_count[0], t=total:
                    self.set_statusbar(f"导入中 ({c}/{t})..."))

        self.root.after(0, lambda: messagebox.showinfo(
            "结果", f"成功: {success[0]} 条\n失败: {fail[0]} 条"))
        self.root.after(0, lambda: self.set_statusbar(
            f"导入完成: 成功 {success[0]}, 失败 {fail[0]}"))

    # ---------- 批量导入文件夹（文件名匹配设备名）----------
    def batch_import_by_folder(self):
        """一键导入：选文件夹，自动匹配文件名与设备名，逐设备导入通道"""
        if not self.access_token:
            messagebox.showwarning("警告", "请先登录")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("缺少库", "请安装 openpyxl: pip install openpyxl")
            return
        if not self.all_devices:
            messagebox.showwarning("提示", "请先刷新设备列表")
            return

        folder = filedialog.askdirectory(title="选择包含 Excel 文件的文件夹")
        if not folder:
            return

        # 扫描 xlsx 文件
        xlsx_files = [os.path.join(folder, f) for f in os.listdir(folder)
                      if f.lower().endswith(".xlsx")]
        if not xlsx_files:
            messagebox.showwarning("无文件", f"文件夹中未找到 .xlsx 文件:\n{folder}")
            return

        # 构建文件名→设备的映射：先用精确名称匹配，再回退到 deviceId 匹配
        matched = []  # [(filepath, device_dict), ...]
        unmatched_files = []

        # 建立快速查找表
        name_to_dev = {}
        id_to_dev = {}
        for dev in self.all_devices:
            dn = dev.get("name", "").strip()
            di = dev.get("deviceId", "").strip()
            if dn:
                name_to_dev[dn] = dev
            if di:
                id_to_dev[di] = dev

        for fpath in xlsx_files:
            fname = os.path.splitext(os.path.basename(fpath))[0].strip()
            if fname in name_to_dev:
                matched.append((fpath, name_to_dev[fname]))
            elif fname in id_to_dev:
                matched.append((fpath, id_to_dev[fname]))
            else:
                unmatched_files.append(os.path.basename(fpath))

        if not matched:
            msg = "未找到文件名与设备名称匹配的 Excel 文件。"
            if unmatched_files:
                msg += f"\n\n文件夹中的文件:\n" + "\n".join(unmatched_files[:15])
            messagebox.showwarning("无匹配", msg)
            return

        # 确认对话框
        msg_lines = [f"找到 {len(matched)} 个匹配，共 {sum(1 for _ in xlsx_files)} 个文件："]
        for fpath, dev in matched:
            msg_lines.append(f"  📄 {os.path.basename(fpath)} → {dev.get('name', '')} ({dev.get('deviceId', '')})")
        if unmatched_files:
            msg_lines.append(f"\n⚠️ {len(unmatched_files)} 个文件未匹配（将被跳过）")
            for fn in unmatched_files[:8]:
                msg_lines.append(f"   - {fn}")
        msg_lines.append("\n确认开始逐设备导入？")

        if not messagebox.askyesno("批量导入确认", "\n".join(msg_lines)):
            return

        self.set_statusbar(f"正在批量导入 {len(matched)} 个设备...")
        threading.Thread(target=self._do_batch_import_by_folder,
                         args=(matched,), daemon=True).start()

    def _do_batch_import_by_folder(self, matched):
        """后台执行：逐设备查通道 → 读 Excel → 匹配 → 提交更新"""
        host = self.server_host.get().strip().rstrip('/')
        token = self.access_token
        token_headers = {"Accept": "*/*", "access-token": token}
        post_headers = {"Accept": "*/*", "access-token": token,
                        "Content-Type": "application/json"}

        total_devices = len(matched)
        total_success = 0
        total_fail = 0
        device_results = []  # [(dev_name, ok_count, fail_count), ...]

        for idx, (fpath, dev) in enumerate(matched, 1):
            dev_name = dev.get("name", "").strip() or dev.get("deviceId", "")
            device_id = dev.get("deviceId", "")
            self.root.after(0, lambda c=idx, t=total_devices, d=dev_name:
                self.set_statusbar(f"批量导入 ({c}/{t}): {d}"))

            # 1. 查询该设备的全部通道
            try:
                channels, total_ch = self._concurrent_channel_query(device_id)
                if not channels:
                    self.root.after(0, lambda d=dev_name:
                        messagebox.showwarning("跳过", f"设备 [{d}] 无通道数据"))
                    continue
            except Exception as e:
                self.root.after(0, lambda d=dev_name, e=e:
                    messagebox.showerror("查询失败", f"设备 [{d}] 通道查询失败:\n{e}"))
                total_fail += 1
                continue

            # 2. 合并全局通道数据（厂家/经纬度/型号等）
            try:
                def merge_one(dc):
                    dc_id = dc.get("id") or dc.get("gbId")
                    if not dc_id:
                        return dc
                    try:
                        one = requests.get(f"{host}/api/common/channel/one",
                                           headers=token_headers,
                                           params={"id": dc_id}, timeout=15)
                        if one.status_code == 200:
                            od = one.json()
                            if od.get("code") == 0:
                                gc = od.get("data")
                                if gc and isinstance(gc, dict):
                                    for fld in ("gbManufacturer", "gbLongitude", "gbLatitude",
                                                "gbName", "gbCivilCode",
                                                "gbModel", "gbOwner", "gbAddress", "gbPassword", "gbDeviceId"):
                                        if fld in gc and gc[fld] is not None:
                                            dc[fld] = gc[fld]
                    except Exception:
                        pass
                    return dc
                with ThreadPoolExecutor(max_workers=self.max_workers) as ex:
                    list(ex.map(merge_one, channels))
            except Exception:
                pass  # 合并失败不影响主流程

            # 3. 读取 Excel 文件
            try:
                wb = load_workbook(fpath)
                ws = wb.active
            except Exception as e:
                self.root.after(0, lambda d=dev_name, f=fpath, e=e:
                    messagebox.showerror("读取失败", f"设备 [{d}] 文件:\n{os.path.basename(f)}\n{e}"))
                total_fail += 1
                continue

            # 4. 解析 Excel → 按数据库ID索引
            excel_data = {}
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 9:
                    continue
                db_id = row[8]
                if not db_id:
                    continue
                try:
                    db_id = int(db_id)
                except (ValueError, TypeError):
                    continue
                excel_data[db_id] = {
                    "name": str(row[2] or "").strip(),
                    "civilCode": str(row[4] or "").strip(),
                    "gbLongitude": str(row[5] or "").strip(),
                    "gbLatitude": str(row[6] or "").strip(),
                    "gbManufacturer": str(row[7] or "").strip(),
                    "gbModel": str(row[9] if len(row) > 9 else "" or "").strip(),
                    "gbOwner": str(row[10] if len(row) > 10 else "" or "").strip(),
                    "gbAddress": str(row[11] if len(row) > 11 else "" or "").strip(),
                    "gbPassword": str(row[12] if len(row) > 12 else "" or "").strip(),
                }

            if not excel_data:
                self.root.after(0, lambda d=dev_name:
                    messagebox.showwarning("无数据", f"设备 [{d}] 的 Excel 中未找到有效数据"))
                total_fail += 1
                continue

            # 5. 匹配通道并构建更新任务
            tasks = []
            for ch in channels:
                cid = ch.get("id")
                if cid in excel_data:
                    d = excel_data[cid]
                    n = d["name"]
                    c = d["civilCode"]
                    lo = d["gbLongitude"]
                    la = d["gbLatitude"]
                    mf = d["gbManufacturer"]
                    md = d.get("gbModel", "")
                    ow = d.get("gbOwner", "")
                    ad = d.get("gbAddress", "")
                    pw = d.get("gbPassword", "")
                    if ((n and n != ch.get("name", "")) or (c and c != ch.get("civilCode", "")) or
                        (lo and str(lo) != str(ch.get("gbLongitude", 0))) or
                        (la and str(la) != str(ch.get("gbLatitude", 0))) or
                        mf != ch.get("gbManufacturer", "") or
                        (md and md != ch.get("gbModel", "")) or
                        (ow and ow != ch.get("gbOwner", "")) or
                        (ad and ad != ch.get("gbAddress", "")) or
                        (pw and pw != ch.get("gbPassword", ""))):
                        tasks.append((ch, n, c, lo, la, mf, md, ow, ad, pw))

            if not tasks:
                device_results.append((dev_name, 0, 0))
                continue

            # 6. 并发提交更新
            dev_ok = [0]
            dev_fail = [0]

            def do_one(task):
                ch, nn, nc, lo, la, mf, md, ow, ad, pw = task
                upd = {}
                if nn and nn != ch.get("gbName", ch.get("name", "")):
                    upd["gbName"] = nn
                if nc and nc != ch.get("gbCivilCode", ch.get("civilCode", "")):
                    upd["gbCivilCode"] = nc
                if lo:
                    try:
                        val = float(lo)
                        if abs(val - float(ch.get("gbLongitude", 0) or 0)) > 0.000001:
                            upd["gbLongitude"] = val
                    except ValueError:
                        pass
                if la:
                    try:
                        val = float(la)
                        if abs(val - float(ch.get("gbLatitude", 0) or 0)) > 0.000001:
                            upd["gbLatitude"] = val
                    except ValueError:
                        pass
                if mf and mf != ch.get("gbManufacturer", ""):
                    upd["gbManufacturer"] = mf
                if md and md != ch.get("gbModel", ""):
                    upd["gbModel"] = md
                if ow and ow != ch.get("gbOwner", ""):
                    upd["gbOwner"] = ow
                if ad and ad != ch.get("gbAddress", ""):
                    upd["gbAddress"] = ad
                if pw and pw != ch.get("gbPassword", ""):
                    upd["gbPassword"] = pw
                if not upd:
                    return True
                try:
                    body = self.build_channel_body(ch, upd)
                    resp = requests.post(f"{host}/api/common/channel/update",
                                         headers=post_headers, json=body, timeout=15)
                    if resp.status_code == 200 and resp.json().get("code") == 0:
                        for k, v in upd.items():
                            ch[k] = v
                        if "gbName" in upd:
                            ch["name"] = upd["gbName"]
                        if "gbCivilCode" in upd:
                            ch["civilCode"] = upd["gbCivilCode"]
                        return True
                except Exception:
                    pass
                return False

            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                futures = {executor.submit(do_one, t): t for t in tasks}
                for future in as_completed(futures):
                    if future.result():
                        dev_ok[0] += 1
                    else:
                        dev_fail[0] += 1

            device_results.append((dev_name, dev_ok[0], dev_fail[0]))
            total_success += dev_ok[0]
            total_fail += dev_fail[0]
            self.root.after(0, lambda d=dev_name, ok=dev_ok[0], fl=dev_fail[0]:
                self.set_statusbar(
                    f"已导入 [{d}]: 成功 {ok}, 失败 {fl}"))

        # 7. 汇总结果
        result_lines = [f"批量导入完成（共 {total_devices} 个设备）",
                        f"总成功: {total_success}  总失败: {total_fail}\n"]
        for dn, ok, fl in device_results:
            result_lines.append(f"  {dn}: ✅ {ok} / ❌ {fl}")
        self.root.after(0, lambda: messagebox.showinfo(
            "批量导入结果", "\n".join(result_lines)))
        self.root.after(0, lambda: self.set_statusbar(
            f"批量导入完成: 成功 {total_success}, 失败 {total_fail}"))

    # ---------- 导入模板下载 ----------
    def download_import_template(self):
        """下载导入模板，标注必填字段和可修改字段"""
        if not HAS_OPENPYXL:
            messagebox.showerror("缺少库", "请安装 openpyxl: pip install openpyxl")
            return

        file = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="保存导入模板",
            initialfile="WVP导入模板.xlsx")
        if not file:
            return

        try:
            wb = Workbook()

            # ---- Sheet 1: 导入模板 ----
            ws = wb.active
            ws.title = "导入模板"

            # 颜色定义
            REQUIRED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")   # 粉红=必填
            EDITABLE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # 绿色=可修改
            INFO_FILL     = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")   # 蓝色=参考
            WHITE_FONT    = Font(color="9C0006", bold=True)
            GREEN_FONT    = Font(color="006100", bold=True)
            BLUE_FONT     = Font(color="1F4E79", bold=True)
            CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 表头定义: (列名, 宽度, fill, font, 备注)
            columns = [
                ("设备名称",   18, INFO_FILL,     BLUE_FONT,  "ⓘ 参考信息，不用于匹配"),
                ("设备ID",     24, INFO_FILL,     BLUE_FONT,  "ⓘ 参考信息"),
                ("通道名称",   25, EDITABLE_FILL, GREEN_FONT, "✎ 可修改"),
                ("通道类型",   10, INFO_FILL,     BLUE_FONT,  "ⓘ 参考信息"),
                ("区域编码",   15, EDITABLE_FILL, GREEN_FONT, "✎ 可修改"),
                ("经度",       12, EDITABLE_FILL, GREEN_FONT, "✎ 可修改"),
                ("纬度",       12, EDITABLE_FILL, GREEN_FONT, "✎ 可修改"),
                ("厂家",       18, EDITABLE_FILL, GREEN_FONT, "✎ 可修改"),
                ("数据库ID",   14, REQUIRED_FILL, WHITE_FONT, "★ 必填——用于匹配通道的唯一标识"),
                ("设备型号",   12, EDITABLE_FILL, GREEN_FONT, "✎ 可修改"),
                ("设备归属",   14, EDITABLE_FILL, GREEN_FONT, "✎ 可修改"),
                ("地址",       20, EDITABLE_FILL, GREEN_FONT, "✎ 可修改"),
                ("密码",       12, EDITABLE_FILL, GREEN_FONT, "✎ 可修改"),
                ("国标编码",   22, INFO_FILL,     BLUE_FONT,  "ⓘ 参考信息"),
            ]

            # 写入表头（第1行）
            for c_idx, (name, width, fill, font, note) in enumerate(columns, 1):
                cell = ws.cell(row=1, column=c_idx, value=name)
                cell.fill = fill
                cell.font = font
                cell.alignment = CENTER
                ws.column_dimensions[cell.column_letter].width = width

            # 写入备注行（第2行）
            for c_idx, (name, width, fill, font, note) in enumerate(columns, 1):
                cell = ws.cell(row=2, column=c_idx, value=note)
                cell.font = Font(color="666666", italic=True, size=8)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            # 冻结表头
            ws.freeze_panes = "A3"

            # ---- Sheet 2: 使用说明 ----
            ws2 = wb.create_sheet("使用说明")
            ws2.column_dimensions['A'].width = 80

            usage = [
                ("WVP 通道导入模板 — 使用说明", True),
                ("", False),
                ("▌ 颜色图例", True),
                ("", False),
                ("  粉红色表头（★）= 必填字段：数据库ID", False),
                ("      用于匹配通道的唯一标识，不可为空，必须与导出时一致。", False),
                ("", False),
                ("  绿色表头（✎）= 可修改字段：", False),
                ("      通道名称、区域编码、经度、纬度、厂家、设备型号、设备归属、地址、密码", False),
                ("      只修改需要变更的单元格，留空的单元格不会被覆盖。", False),
                ("", False),
                ("  蓝色表头（ⓘ）= 参考信息字段：", False),
                ("      设备名称、设备ID、通道类型、国标编码", False),
                ("      仅供参考，修改这些列不会影响导入结果。", False),
                ("", False),
                ("▌ 使用步骤", True),
                ("", False),
                ("  1. 在工具中双击设备 → 点击「导出Excel」", False),
                ("       → 得到带实际数据的完整 Excel", False),
                ("  2. 在此 Excel 中修改需要变更的单元格", False),
                ("      保留「数据库ID」列不变，这是匹配依据", False),
                ("  3. 在工具中点击「导入通道Excel」或「一键导入文件夹」", False),
                ("       → 选择修改后的 Excel → 确认提交", False),
                ("", False),
                ("  或直接从零开始：", False),
                ("  在此模板 Sheet 中按格式填入数据即可。", False),
                ("", False),
                ("▌ 一键导入文件夹（按设备名匹配）", True),
                ("", False),
                ("  将每个设备的导出 Excel 以其设备名命名", False),
                ("  例：摄像头01.xlsx、摄像头02.xlsx", False),
                ("  点击「一键导入文件夹」选中存放这些文件的文件夹", False),
                ("  工具会自动按文件名匹配设备名并逐设备导入。", False),
                ("", False),
                ("▌ 注意事项", True),
                ("", False),
                ("  • 经度/纬度必须是数字（如 120.5），留空则不修改", False),
                ("  • 数据库ID列不可为空，每行必须有值", False),
                ("  • 如果数据库ID不在通道列表中，该行会被跳过", False),
                ("  • 建议先导出 → 修改 → 导入，确保ID一致", False),
            ]

            for i, (text, is_title) in enumerate(usage, 1):
                cell = ws2.cell(row=i, column=1, value=text)
                if is_title:
                    cell.font = Font(bold=True, size=12, color="1F4E79")
                else:
                    cell.font = Font(size=10)

            wb.save(file)
            self.set_statusbar(f"模板已保存: {os.path.basename(file)}")
            messagebox.showinfo("成功", f"导入模板已保存到:\n{file}\n\n包含两个 Sheet:\n  • 导入模板 — 带颜色标注的字段\n  • 使用说明 — 操作指南")
        except Exception as e:
            messagebox.showerror("失败", str(e))

    # ---------- 导出Excel ----------
    def export_excel(self):
        if not self.all_channels:
            messagebox.showwarning("提示", "通道列表为空")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("缺少库", "请安装 openpyxl: pip install openpyxl")
            return
        default_name = f"{self.selected_device_name or '通道列表'}.xlsx"
        file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx")],
                                             title="导出通道列表",
                                             initialfile=default_name)
        if not file:
            return
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "通道列表"
            headers = ["序号", "设备ID", "名称", "通道类型", "区域编码", "经度", "纬度", "厂家", "数据库ID",
                       "设备型号", "设备归属", "地址", "密码", "国标编码"]
            for c, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=c, value=h)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")
            for i, ch in enumerate(self.all_channels, 1):
                row = [i, ch.get("deviceId", ""), ch.get("name", ""),
                       "子目录" if ch.get("channelType") else "设备通道",
                       ch.get("civilCode", ""), ch.get("gbLongitude", 0), ch.get("gbLatitude", 0), ch.get("gbManufacturer", ""), ch.get("id", ""),
                       ch.get("gbModel", ""), ch.get("gbOwner", ""),
                       ch.get("gbAddress", ""), ch.get("gbPassword", ""),
                       ch.get("gbDeviceId", "")]
                for j, v in enumerate(row, 1):
                    ws.cell(row=i+1, column=j, value=v)
            for i, w in enumerate([6, 22, 25, 10, 15, 10, 10, 15, 10, 10, 12, 18, 10, 10], 1):
                ws.column_dimensions[ws.cell(1, i).column_letter].width = w
            wb.save(file)
            self.set_statusbar(f"导出成功: {os.path.basename(file)}")
            messagebox.showinfo("成功", f"已导出到:\n{file}")
        except Exception as e:
            messagebox.showerror("失败", str(e))

    # ---------- 导入Excel（带进度条）----------
    def import_excel(self):
        if not self.access_token:
            messagebox.showwarning("警告", "请先登录")
            return
        if not HAS_OPENPYXL:
            messagebox.showerror("缺少库", "请安装 openpyxl: pip install openpyxl")
            return
        if not self.all_channels:
            messagebox.showwarning("提示", "请先查询通道再导入")
            return
        file = filedialog.askopenfilename(title="选择修改后的Excel", filetypes=[("Excel", "*.xlsx")])
        if not file:
            return
        try:
            wb = load_workbook(file)
            ws = wb.active
        except Exception as e:
            messagebox.showerror("读取失败", str(e))
            return
        excel = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) < 9:
                continue
            did = row[8]
            if not did:
                continue
            try:
                did = int(did)
            except:
                continue
            excel[did] = {"name": str(row[2] or "").strip(), "civilCode": str(row[4] or "").strip(), "gbLongitude": str(row[5] or "").strip(), "gbLatitude": str(row[6] or "").strip(), "gbManufacturer": str(row[7] or "").strip(),
                       "gbModel": str(row[9] if len(row) > 9 else "" or "").strip(),
                       "gbOwner": str(row[10] if len(row) > 10 else "" or "").strip(),
                       "gbAddress": str(row[11] if len(row) > 11 else "" or "").strip(),
                       "gbPassword": str(row[12] if len(row) > 12 else "" or "").strip()}
        if not excel:
            messagebox.showwarning("无数据", "Excel无有效数据")
            return
        tasks = []
        for ch in self.all_channels:
            if ch["id"] in excel:
                n = excel[ch["id"]]["name"]
                c = excel[ch["id"]]["civilCode"]
                lo = excel[ch["id"]]["gbLongitude"]
                la = excel[ch["id"]]["gbLatitude"]
                mf = excel[ch["id"]]["gbManufacturer"]
                md = excel[ch["id"]].get("gbModel", "")
                ow = excel[ch["id"]].get("gbOwner", "")
                ad = excel[ch["id"]].get("gbAddress", "")
                pw = excel[ch["id"]].get("gbPassword", "")
                if ((n and n != ch.get("name", "")) or (c and c != ch.get("civilCode", "")) or
                    (lo and str(lo) != str(ch.get("gbLongitude", 0))) or
                    (la and str(la) != str(ch.get("gbLatitude", 0))) or
                    mf != ch.get("gbManufacturer", "") or
                    (md and md != ch.get("gbModel", "")) or
                    (ow and ow != ch.get("gbOwner", "")) or
                    (ad and ad != ch.get("gbAddress", "")) or
                    (pw and pw != ch.get("gbPassword", ""))):
                    tasks.append((ch, n, c, lo, la, mf, md, ow, ad, pw))
        if not tasks:
            messagebox.showinfo("提示", "没有需要修改的数据")
            return
        if not messagebox.askyesno("确认", f"检测到 {len(tasks)} 条修改，是否继续?"):
            return
        self.progress = ProgressDialog(self.root, "正在导入修改", len(tasks))
        self.import_btn.configure(state=tk.DISABLED)
        threading.Thread(target=self.batch_excel_update, args=(tasks,), daemon=True).start()

    def batch_excel_update(self, tasks):
        success = fail = 0
        host = self.server_host.get().strip().rstrip('/')
        headers = {"Accept": "*/*", "access-token": self.access_token, "Content-Type": "application/json"}
        total = len(tasks)
        for i, (ch, nn, nc, lo, la, mf, md, ow, ad, pw) in enumerate(tasks, 1):
            if self.progress and self.progress.is_cancelled():
                break
            upd = {}
            if nn and nn != ch.get("gbName", ch.get("name", "")):
                upd["gbName"] = nn
            if nc and nc != ch.get("gbCivilCode", ch.get("civilCode", "")):
                upd["gbCivilCode"] = nc
                upd["gbCivilCode"] = nc
            if lo and str(lo) != str(ch.get("gbLongitude", 0)):
                try:
                    val = float(lo)
                    upd["gbLongitude"] = val
                except ValueError:
                    pass
            if la and str(la) != str(ch.get("gbLatitude", 0)):
                try:
                    val = float(la)
                    upd["gbLatitude"] = val
                except ValueError:
                    pass
            if mf != ch.get("gbManufacturer", ""):
                upd["gbManufacturer"] = mf
            if md and md != ch.get("gbModel", ""):
                upd["gbModel"] = md
            if ow and ow != ch.get("gbOwner", ""):
                upd["gbOwner"] = ow
            if ad and ad != ch.get("gbAddress", ""):
                upd["gbAddress"] = ad
            if pw and pw != ch.get("gbPassword", ""):
                upd["gbPassword"] = pw
            try:
                body = self.build_channel_body(ch, upd)
                resp = requests.post(f"{host}/api/common/channel/update", headers=headers, json=body, timeout=15)
                if resp.status_code == 200 and resp.json().get("code") == 0:
                    if "gbName" in upd:
                        ch["name"] = upd["gbName"]
                    if "gbCivilCode" in upd:
                        ch["civilCode"] = upd["gbCivilCode"]
                    success += 1
                else:
                    fail += 1
            except:
                fail += 1
            self.root.after(0, lambda c=i, t=total: self.progress.update(c, t, f"正在处理 {c}/{t}"))
        self.root.after(0, self.progress.close)
        self.root.after(0, lambda: self._batch_finished(success, fail))
        self.root.after(0, lambda: self.import_btn.configure(state=tk.NORMAL))

    # ---------- 退出 ----------
    def logout(self):
        if messagebox.askyesno("退出", "确定要退出登录吗？"):
            self.root.destroy()

    def on_close(self):
        self.cancel_edit()
        self.root.destroy()
